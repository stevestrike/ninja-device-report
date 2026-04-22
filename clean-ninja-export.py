#!/usr/bin/env python3
"""
Clean a NinjaRMM device export CSV into a client-presentable XLSX.

Usage:
    python clean_ninja_export.py <input.csv> [output.xlsx]

If output path is omitted, writes to <input>_cleaned.xlsx alongside the input.
"""
import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


KEEP_COLUMNS = [
    "Organization",
    "Location",
    "Display Name",
    "Type",
    "Device Role",
    "Policy",
    "Last Update",
    "Warranty Start Date_formatted",
    "Warranty End Date_formatted",
    "Last Login",
    "Memory Capacity GiB",
    "OS Name",
    "System Name",
    "Device Model",
    "Serial Number",
    "Processors Name",
    "Volumes",
]


# --- Processor cleanup -------------------------------------------------------

def clean_processor(raw):
    """Turn Ninja's CPU strings into short client-readable form.

    Handles dedup + socket count for comma-separated multi-CPU entries.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    # Ninja joins multi-socket CPUs with commas; sometimes with trailing spaces.
    parts = [p.strip() for p in s.split(",") if p.strip()]
    cleaned_parts = [_clean_single_cpu(p) for p in parts]

    # Collapse duplicates while tracking count.
    if not cleaned_parts:
        return ""

    # All identical -> count prefix
    first = cleaned_parts[0]
    if all(p == first for p in cleaned_parts):
        count = len(cleaned_parts)
        return f"{count}\u00d7 {first}" if count > 1 else first

    # Mixed CPUs (rare) -> comma-join uniques preserving order
    seen = []
    for p in cleaned_parts:
        if p not in seen:
            seen.append(p)
    return ", ".join(seen)


def _clean_single_cpu(s):
    # Remove (R) and (TM) trademark markers
    s = re.sub(r"\([Rr]\)", "", s)
    s = re.sub(r"\([Tt][Mm]\)", "", s)

    # Drop the word "CPU" (whole word)
    s = re.sub(r"\bCPU\b", "", s)

    # Drop trailing clock spec like "@ 2.40GHz" or "@ 2.80 GHz"
    s = re.sub(r"@\s*\d+(?:\.\d+)?\s*[GM]Hz", "", s, flags=re.IGNORECASE)

    # Drop AMD graphics trailer, e.g. "with Radeon Graphics" or "w/ Radeon Graphics"
    s = re.sub(r"\s+(?:with|w/)\s+Radeon(?:\s+\w+)?\s+Graphics", "", s, flags=re.IGNORECASE)

    # Drop "Processor" word some AMD strings use
    s = re.sub(r"\bProcessor\b", "", s, flags=re.IGNORECASE)

    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    # Strip stray trailing punctuation
    s = s.rstrip(",; ")
    return s


# --- Volumes cleanup ---------------------------------------------------------

_VOL_KV = re.compile(r'(\w[\w %]*?):\s*"([^"]*)"')

def clean_volumes(raw):
    """Extract the C: drive and format as 'C: X.X / Y.Y GiB free (Z% used)'."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    # Ninja concatenates multiple volumes; split on literal " | " or newline if present,
    # then fall back to scanning by drive Name markers.
    candidates = re.split(r"\s*\|\s*|\r?\n", s)
    if len(candidates) == 1:
        # Try splitting by "Name: " occurrences (keep the delimiter)
        chunks = re.split(r'(?=Name:\s*")', s)
        candidates = [c for c in chunks if c.strip()]

    for chunk in candidates:
        kv = dict(_VOL_KV.findall(chunk))
        name = kv.get("Name", "").strip()
        if name.upper().startswith("C"):
            return _format_volume(kv)

    return ""


def _format_volume(kv):
    cap_raw = kv.get("Capacity", "")
    free_raw = kv.get("Free", "")
    usage = kv.get("Usage %", "").strip().rstrip("%")

    cap_gib = _parse_gib(cap_raw)
    free_gib = _parse_gib(free_raw)

    if cap_gib is None:
        return ""

    # Round free to 1 decimal, capacity to 1 decimal
    cap_s = f"{cap_gib:.1f}"
    free_s = f"{free_gib:.1f}" if free_gib is not None else "?"

    if usage:
        return f"C: {free_s} / {cap_s} GiB free ({usage}% used)"
    return f"C: {free_s} / {cap_s} GiB free"


def _parse_gib(value):
    """Extract the GiB number from a string like '252850466816 (235.5 GiB)'."""
    if not value:
        return None
    m = re.search(r"([\d.]+)\s*GiB", value)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    # Fall back: if only bytes are given, convert
    m = re.search(r"^\s*(\d+)\s*$", value)
    if m:
        try:
            return int(m.group(1)) / (1024 ** 3)
        except ValueError:
            return None
    return None


# --- OS Name cleanup ---------------------------------------------------------

def clean_os_name(raw):
    """Collapse 'Microsoft Windows 11 Pro 10.0.22631' -> 'Win 11 Pro' etc."""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).strip()
    if not s:
        return ""

    # Drop trailing build numbers like "10.0.22631" or "10.0.19045.3803"
    s = re.sub(r"\s+\d+(?:\.\d+){1,3}\s*$", "", s)

    # Drop leading "Microsoft "
    s = re.sub(r"^Microsoft\s+", "", s, flags=re.IGNORECASE)

    # Windows -> Win
    s = re.sub(r"^Windows\b", "Win", s)

    # Collapse whitespace
    s = re.sub(r"\s+", " ", s).strip()
    return s


# --- Memory rounding ---------------------------------------------------------

# Standard RAM capacities clients actually buy, in GB.
# Covers mobile/desktop/workstation/server. Snap to nearest tier within tolerance.
STANDARD_MEMORY_TIERS = [
    1, 2, 4, 6, 8, 12, 16, 24, 32, 48, 64, 96, 128, 192, 256, 384, 512,
    768, 1024, 1536, 2048, 3072, 4096,
]


def clean_memory(raw):
    """Snap reported RAM to the nearest standard purchase tier.

    Ninja reports usable memory, which is slightly less than installed (BIOS
    reserve, integrated graphics, etc.). A 32GB laptop may report 31.42 GiB;
    a 16GB machine may report 15.87 GiB. We round UP to the nearest standard
    tier if the reported value is within 5% below it, otherwise fall back to
    rounding to the nearest tier.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    try:
        val = float(raw)
    except (ValueError, TypeError):
        return raw

    if val <= 0:
        return 0

    # Prefer rounding UP to the nearest tier if within 5% (captures
    # "15.87 -> 16", "31.42 -> 32", "255.5 -> 256" cases).
    for tier in STANDARD_MEMORY_TIERS:
        if val <= tier and val >= tier * 0.95:
            return tier

    # Otherwise snap to whichever tier is closest (handles oddball configs).
    return min(STANDARD_MEMORY_TIERS, key=lambda t: abs(t - val))


# --- Main pipeline -----------------------------------------------------------

def process(input_path: Path, output_path: Path):
    df = pd.read_csv(input_path, dtype=str, keep_default_na=False)

    # Warn on missing expected columns but don't fail — just skip them
    missing = [c for c in KEEP_COLUMNS if c not in df.columns]
    if missing:
        print(f"WARNING: input is missing expected columns: {missing}", file=sys.stderr)

    present_cols = [c for c in KEEP_COLUMNS if c in df.columns]
    out = df[present_cols].copy()

    # Apply cleanups
    if "Processors Name" in out.columns:
        out["Processors Name"] = out["Processors Name"].apply(clean_processor)
    if "Volumes" in out.columns:
        out["Volumes"] = out["Volumes"].apply(clean_volumes)
    if "OS Name" in out.columns:
        out["OS Name"] = out["OS Name"].apply(clean_os_name)
    if "Memory Capacity GiB" in out.columns:
        out["Memory Capacity GiB"] = out["Memory Capacity GiB"].apply(clean_memory)

    _write_xlsx(out, output_path)
    print(f"Wrote {len(out)} rows to {output_path}")


def _write_xlsx(df, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    # Header row
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="BFBFBF")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border

    # Data rows
    body_font = Font(name="Calibri", size=10)
    body_align = Alignment(horizontal="left", vertical="top", wrap_text=False)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            if val == "" or (isinstance(val, float) and pd.isna(val)):
                val = None
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = body_font
            c.alignment = body_align
            c.border = cell_border

    # Freeze header and enable autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Column widths — rough auto-size based on content length, capped
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for val in df[col_name].astype(str):
            if len(val) > max_len:
                max_len = len(val)
        # Cap width to keep it sensible
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input", help="Input NinjaRMM CSV export")
    ap.add_argument("output", nargs="?", help="Output XLSX path (optional)")
    args = ap.parse_args()

    in_path = Path(args.input)
    if not in_path.exists():
        print(f"ERROR: input file not found: {in_path}", file=sys.stderr)
        sys.exit(1)

    if args.output:
        out_path = Path(args.output)
    else:
        out_path = in_path.with_name(in_path.stem + "_cleaned.xlsx")

    process(in_path, out_path)


if __name__ == "__main__":
    main()